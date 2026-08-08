import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# 1. Base de dados
data = [
    ["DORA-1.1", "ICT Risk Management", "ICT Risk Management Framework documented", "Fully Compliant", "Verified", "Low", "GRC & Security"],
    ["DORA-1.2", "ICT Risk Management", "Continuous asset mapping & identification", "Partially Compliant", "In Review", "Medium", "Infra & Sec"],
    ["DORA-1.3", "ICT Risk Management", "Annual review of ICT risk policies", "Fully Compliant", "Verified", "Low", "GRC & Security"],
    ["DORA-1.4", "ICT Risk Management", "Legacy system risk assessment", "Non-Compliant", "Missing", "Critical", "Enterprise Arch"],
    ["DORA-1.5", "ICT Risk Management", "Cryptographic controls & key management", "Partially Compliant", "In Review", "High", "Infra & Sec"],
    ["DORA-1.6", "ICT Risk Management", "Access control & identity management policy", "Fully Compliant", "Verified", "Low", "IAM Team"],
    ["DORA-1.7", "ICT Risk Management", "Multi-factor authentication (MFA) enforcement", "Fully Compliant", "Verified", "Low", "IAM Team"],
    ["DORA-1.8", "ICT Risk Management", "Network segmentation between prod/non-prod", "Partially Compliant", "In Review", "Medium", "Network Sec"],
    ["DORA-1.9", "ICT Risk Management", "Patch management SLA enforcement", "Non-Compliant", "Missing", "High", "Infra & Sec"],
    ["DORA-1.10", "ICT Risk Management", "Data loss prevention (DLP) monitoring", "Non-Compliant", "Missing", "High", "AppSec"],
    ["DORA-2.1", "ICT Incident Management", "ICT Incident classification methodology", "Fully Compliant", "Verified", "Low", "SOC Team"],
    ["DORA-2.2", "ICT Incident Management", "Major incident notification to regulator (<2h)", "Non-Compliant", "Missing", "Critical", "Regulatory Compliance"],
    ["DORA-2.3", "ICT Incident Management", "Centralized incident logging & SIEM integration", "Fully Compliant", "Verified", "Low", "SOC Team"],
    ["DORA-2.4", "ICT Incident Management", "Post-incident root cause analysis process", "Partially Compliant", "In Review", "Medium", "SOC Team"],
    ["DORA-2.5", "ICT Incident Management", "Crisis communication plan with stakeholders", "Fully Compliant", "Verified", "Low", "Corp Comms"],
    ["DORA-2.6", "ICT Incident Management", "Customer notification process for major incidents", "Partially Compliant", "In Review", "High", "Customer Operations"],
    ["DORA-2.7", "ICT Incident Management", "Incident response playbooks for ransomware", "Non-Compliant", "Missing", "Critical", "SOC Team"],
    ["DORA-2.8", "ICT Incident Management", "Anomalous behavior detection mechanisms", "Fully Compliant", "Verified", "Low", "SOC Team"],
    ["DORA-3.1", "Digital Operational Resilience Testing", "Annual vulnerability scanning program", "Fully Compliant", "Verified", "Low", "AppSec"],
    ["DORA-3.2", "Digital Operational Resilience Testing", "Threat-Led Penetration Testing (TLPT)", "Non-Compliant", "Missing", "Critical", "Red Team"],
    ["DORA-3.3", "Digital Operational Resilience Testing", "Independent testing of critical ICT tools", "Partially Compliant", "In Review", "High", "Audit Team"],
    ["DORA-3.4", "Digital Operational Resilience Testing", "Source code security review integration", "Fully Compliant", "Verified", "Low", "DevSecOps"],
    ["DORA-3.5", "Digital Operational Resilience Testing", "Disaster recovery plan dry-run testing", "Partially Compliant", "In Review", "Medium", "BCP Team"],
    ["DORA-3.6", "Digital Operational Resilience Testing", "Backup restoration verification testing", "Non-Compliant", "Missing", "Critical", "Infra & Sec"],
    ["DORA-3.7", "Digital Operational Resilience Testing", "Third-party penetration testing validation", "Partially Compliant", "In Review", "Medium", "TPRM Team"],
    ["DORA-3.8", "Digital Operational Resilience Testing", "Red team scenario exercises", "Non-Compliant", "Missing", "High", "Red Team"],
    ["DORA-4.1", "Third-Party Risk Management", "Register of Information (RoI) for ICT vendors", "Partially Compliant", "In Review", "High", "TPRM Team"],
    ["DORA-4.2", "Third-Party Risk Management", "Critical vendor concentration risk assessment", "Non-Compliant", "Missing", "Critical", "TPRM Team"],
    ["DORA-4.3", "Third-Party Risk Management", "Subcontracting chain risk monitoring", "Non-Compliant", "Missing", "High", "TPRM Team"],
    ["DORA-4.4", "Third-Party Risk Management", "Mandatory DORA clauses in vendor contracts", "Partially Compliant", "In Review", "High", "Legal & Procurement"],
    ["DORA-4.5", "Third-Party Risk Management", "Vendor exit strategy & transition plans", "Non-Compliant", "Missing", "Critical", "TPRM Team"],
    ["DORA-4.6", "Third-Party Risk Management", "Annual performance audit of critical vendors", "Fully Compliant", "Verified", "Low", "TPRM Team"],
    ["DORA-4.7", "Third-Party Risk Management", "SLA monitoring for critical ICT services", "Fully Compliant", "Verified", "Low", "Vendor Management"],
    ["DORA-4.8", "Third-Party Risk Management", "Vendor security questionnaire automation", "Partially Compliant", "In Review", "Medium", "TPRM Team"],
    ["DORA-5.1", "Information Sharing", "Threat intelligence sharing participation", "Fully Compliant", "Verified", "Low", "Threat Intel"],
    ["DORA-5.2", "Information Sharing", "Internal threat intelligence distribution", "Fully Compliant", "Verified", "Low", "Threat Intel"],
    ["DORA-5.3", "Information Sharing", "Peer-to-peer IoC exchange mechanism", "Partially Compliant", "In Review", "Medium", "Threat Intel"],
    ["DORA-5.4", "Information Sharing", "Regulatory cyber threat reporting", "Fully Compliant", "Verified", "Low", "Regulatory Compliance"]
]

cols = ["Requirement_ID", "Domain", "Requirement_Title", "Status", "Evidence_Status", "Risk_Level", "Owner_Team"]
df = pd.DataFrame(data, columns=cols)
df.to_csv('dora_compliance_master.csv', index=False)

# 2. Excel Setup
output_file = "DORA_Compliance_Executive_Tracker.xlsx"
wb = openpyxl.Workbook()
wb.remove(wb.active)

font_name = "Calibri"
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

status_styles = {
    "Fully Compliant": {"fill": "DCFCE7", "font": "15803D"},
    "Partially Compliant": {"fill": "FEF3C7", "font": "B45309"},
    "Non-Compliant": {"fill": "FEE2E2", "font": "B91C1C"}
}

# 3. Executive Summary
ws_dash = wb.create_sheet(title="Executive Summary")
ws_dash.views.sheetView[0].showGridLines = False

ws_dash.merge_cells("A1:E1")
ws_dash["A1"].value = "DORA COMPLIANCE & GAP ANALYSIS DASHBOARD"
ws_dash["A1"].font = Font(name=font_name, size=14, bold=True, color="1E3A8A")

ws_dash.merge_cells("A2:E2")
ws_dash["A2"].value = "EU Regulation (EU) 2022/2554 • Illustrative Benchmarking Model (Fictional Data)"
ws_dash["A2"].font = Font(name=font_name, size=9, italic=True, color="6B7280")

total_reqs = len(df)
fully_comp = len(df[df['Status'] == 'Fully Compliant'])
part_comp = len(df[df['Status'] == 'Partially Compliant'])
non_comp = len(df[df['Status'] == 'Non-Compliant'])
compliance_rate = round((fully_comp / total_reqs) * 100, 1)

ws_dash.cell(row=4, column=1, value="Metric / Indicator")
ws_dash.cell(row=4, column=2, value="Count")

metrics = [
    ("Fully Compliant", fully_comp),
    ("Partially Compliant", part_comp),
    ("Non-Compliant", non_comp),
    ("Total Evaluated", total_reqs)
]

for idx, (label, val) in enumerate(metrics, start=5):
    ws_dash.cell(row=idx, column=1, value=label)
    ws_dash.cell(row=idx, column=2, value=val)

ws_dash.cell(row=9, column=1, value="Overall Compliance Rate")
ws_dash.cell(row=9, column=2, value=f"{compliance_rate}%")

for r in range(4, 10):
    for c in range(1, 3):
        cell = ws_dash.cell(row=r, column=c)
        cell.border = thin_border
        if r == 4:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if c == 2 else "left")
        else:
            cell.font = Font(name=font_name, size=10, bold=(r == 9 or c == 1))
            if c == 2:
                cell.alignment = Alignment(horizontal="center")

ws_dash.cell(row=12, column=1, value="Domain Summary")
ws_dash.cell(row=12, column=2, value="Compliant")
ws_dash.cell(row=12, column=3, value="Gaps")
ws_dash.cell(row=12, column=4, value="Total")
ws_dash.cell(row=12, column=5, value="% Compliance")

for c in range(1, 6):
    cell = ws_dash.cell(row=12, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center" if c > 1 else "left")

curr_r = 13
for dom in df['Domain'].unique():
    dom_sub = df[df['Domain'] == dom]
    d_tot = len(dom_sub)
    d_comp = len(dom_sub[dom_sub['Status'] == 'Fully Compliant'])
    d_gaps = len(dom_sub[dom_sub['Status'] != 'Fully Compliant'])
    d_rate = round((d_comp / d_tot) * 100, 1) / 100
    
    ws_dash.cell(row=curr_r, column=1, value=dom)
    ws_dash.cell(row=curr_r, column=2, value=d_comp)
    ws_dash.cell(row=curr_r, column=3, value=d_gaps)
    ws_dash.cell(row=curr_r, column=4, value=d_tot)
    
    p_cell = ws_dash.cell(row=curr_r, column=5, value=d_rate)
    p_cell.number_format = '0.0%'
    
    for c in range(1, 6):
        cell = ws_dash.cell(row=curr_r, column=c)
        cell.border = thin_border
        cell.font = Font(name=font_name, size=10)
        if c > 1:
            cell.alignment = Alignment(horizontal="center")
    curr_r += 1

disc_row = curr_r + 2
ws_dash.merge_cells(start_row=disc_row, start_column=1, end_row=disc_row, end_column=5)
ws_dash.cell(row=disc_row, column=1, value="Note: Portfolio demonstration created in compliance with Regulation (EU) 2022/2554 (DORA). All data points are mock values.").font = Font(name=font_name, size=8, italic=True, color="9CA3AF")

# 4. Charts
pie = PieChart()
pie.title = "Status Distribution"
pie.title.text.font = Font(name=font_name, size=11, bold=True)
pie.add_data(Reference(ws_dash, min_col=2, min_row=4, max_row=7), titles_from_data=True)
pie.set_categories(Reference(ws_dash, min_col=1, min_row=5, max_row=7))
pie.width, pie.height = 13, 7
ws_dash.add_chart(pie, "G4")

bar = BarChart()
bar.type = "col"
bar.style = 10
bar.title = "Compliance Level by DORA Domain"
bar.y_axis.title, bar.x_axis.title = "Requirements", "Domain"
bar.width, bar.height = 15, 7.5
bar.add_data(Reference(ws_dash, min_col=2, max_col=3, min_row=12, max_row=curr_r-1), titles_from_data=True)
bar.set_categories(Reference(ws_dash, min_col=1, min_row=13, max_row=curr_r-1))
ws_dash.add_chart(bar, "G18")

# 5. All Requirements Tab
ws_all = wb.create_sheet(title="All Requirements")
ws_all.views.sheetView[0].showGridLines = False

headers = list(df.columns)
ws_all.append(headers)

for c_idx in range(1, len(headers) + 1):
    cell = ws_all.cell(row=1, column=c_idx)
    cell.fill, cell.font = header_fill, header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for _, row in df.iterrows():
    row_vals = list(row)
    ws_all.append(row_vals)
    c_row = ws_all.max_row
    
    for c_idx in range(1, len(row_vals) + 1):
        cell = ws_all.cell(row=c_row, column=c_idx)
        cell.border = thin_border
        cell.font = Font(name=font_name, size=10)
        
        if headers[c_idx - 1] == "Status":
            s_val = str(cell.value)
            if s_val in status_styles:
                cfg = status_styles[s_val]
                cell.fill = PatternFill(start_color=cfg["fill"], end_color=cfg["fill"], fill_type="solid")
                cell.font = Font(name=font_name, size=10, bold=True, color=cfg["font"])
            cell.alignment = Alignment(horizontal="center")
            
        if headers[c_idx - 1] in ["Requirement_ID", "Risk_Level", "Evidence_Status"]:
            cell.alignment = Alignment(horizontal="center")

# 6. Urgent Gaps Tab
ws_gaps = wb.create_sheet(title="Urgent Remediation Gaps")
ws_gaps.views.sheetView[0].showGridLines = False

urgent_gaps = df[(df['Status'] == 'Non-Compliant') & (df['Risk_Level'].isin(['Critical', 'High']))]
ws_gaps.append(headers)

red_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
for c_idx in range(1, len(headers) + 1):
    cell = ws_gaps.cell(row=1, column=c_idx)
    cell.fill, cell.font = red_fill, header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for _, row in urgent_gaps.iterrows():
    row_vals = list(row)
    ws_gaps.append(row_vals)
    c_row = ws_gaps.max_row
    
    for c_idx in range(1, len(row_vals) + 1):
        cell = ws_gaps.cell(row=c_row, column=c_idx)
        cell.border = thin_border
        cell.font = Font(name=font_name, size=10)
        
        if headers[c_idx - 1] == "Status":
            cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            cell.font = Font(name=font_family if 'font_family' in locals() else font_name, size=10, bold=True, color="B91C1C")
            cell.alignment = Alignment(horizontal="center")
            
        if headers[c_idx - 1] in ["Requirement_ID", "Risk_Level", "Evidence_Status"]:
            cell.alignment = Alignment(horizontal="center")

# 7. Auto-fit columns & save
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

ws_dash.column_dimensions['F'].width = 4
ws_dash.column_dimensions['G'].width = 18

wb.save(output_file)
print("Done.")